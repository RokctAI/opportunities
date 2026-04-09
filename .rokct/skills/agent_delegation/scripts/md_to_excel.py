# Licensed under the MIT License.
# Copyright 2024 RokctAI

import os
import pandas as pd
import re
from pathlib import Path
from datetime import datetime
from dotenv import load_dotenv

def load_environment():
    """Robust environment loading from .env/production.env."""
    script_dir = os.path.dirname(os.path.abspath(__file__))
    project_root = os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(script_dir))))
    env_path = os.path.join(project_root, ".env", "production.env")

    if os.path.exists(env_path):
        load_dotenv(env_path)
        with open(env_path, 'r', encoding='utf-8') as f:
            for line in f:
                if "=" in line and not line.startswith("#"):
                    key, val = line.replace("export ", "").strip().split("=", 1)
                    if not os.environ.get(key.strip()):
                        os.environ[key.strip()] = val.strip("'\" ")
    else:
        load_dotenv()
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import landscape, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet

# --- GLOBAL THEMES ---
INDIGO_HEADER = "4B0082"
EMERALD_HEADER = "50C878"
RUBY_HEADER = "E0115F"  # For Tenders
TEXT_WHITE = "FFFFFF"
STRIPE_GRAY = "F2F2F2"

def apply_premium_style(writer, sheet_name, header_color):
    workbook = writer.book
    worksheet = writer.sheets[sheet_name]
    
    header_fill = PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")
    header_font = Font(bold=True, color=TEXT_WHITE)
    stripe_fill = PatternFill(start_color=STRIPE_GRAY, end_color=STRIPE_GRAY, fill_type="solid")
    
    # 1. Style Header
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # 2. Zebra Stripes & Filters
    rows = list(worksheet.rows)
    for i, row in enumerate(rows[1:], start=2):
        if i % 2 == 0:
            for cell in row:
                cell.fill = stripe_fill
    
    # 3. Freeze Top Row
    worksheet.freeze_panes = "A2"
    
    # 4. Auto-Filter
    worksheet.auto_filter.ref = worksheet.dimensions

    # 5. Auto-Column Width & Wrapping
    for col_idx in range(1, worksheet.max_column + 1):
        max_length = 0
        column = get_column_letter(col_idx)
        is_multiline = False
        for cell in worksheet[column]:
            if cell.row == 1: continue # Skip header for length
            try:
                val = str(cell.value)
                if '\n' in val: is_multiline = True
                if len(val) > max_length:
                    max_length = len(val)
            except:
                pass

        # Enable wrapping for multiline cells (like descriptions and multi-links)
        if is_multiline:
            for cell in worksheet[column]:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            adjusted_width = 40
        else:
            adjusted_width = min(max_length + 2, 50)

        worksheet.column_dimensions[column].width = adjusted_width

def parse_markdown_table(md_content):
    lines = md_content.split('\n')
    table_lines = [l.strip() for l in lines if l.strip().startswith('|')]
    if len(table_lines) < 2: return None
    
    # Skip separator row
    clean_lines = [l for l in table_lines if not re.match(r'^\|\s*[:\-|\s]+\s*\|$', l)]
    
    rows = []
    for line in clean_lines:
        cells = [cell.strip() for cell in line.strip('|').split('|')]
        rows.append(cells)
    
    if len(rows) < 2: return None
    return pd.DataFrame(rows[1:], columns=rows[0])

def parse_discrete_markdown(md_content):
    data = {}
    title_match = re.search(r'^#\s+(.+)$', md_content, re.MULTILINE)
    data['Grant Title'] = title_match.group(1).strip() if title_match else "Unknown"
    
    # Extract Description from the section below ## Description
    desc_match = re.search(r'## Description\s*\n\s*([\s\S]+?)\n\n##', md_content)
    if not desc_match:
        desc_match = re.search(r'## Description\s*\n\s*([\s\S]+?)\n\n', md_content)
    data['Description'] = desc_match.group(1).strip() if desc_match else "No description provided."

    fields = {
        'Organization': r'-\s+\*\*Organization\*\*:\s*(.+)$',
        'Deadline': r'-\s+\*\*Deadline\*\*:\s*(.+)$',
        'Amount': r'-\s+\*\*Funding Amount\*\*:\s*(.+)$',
        'Focus Area': r'-\s+\*\*Focus Area\*\*:\s*(.+)$',
        'Apply Link': r'-\s+\*\*Applying Link\*\*:\s*(.+)$',
        'Source': r'-\s+\*\*Source\*\*:\s*(.+)$',
        'Status': r'-\s+\*\*Verification Status\*\*:\s*(.+)$'
    }
    for label, pattern in fields.items():
        match = re.search(pattern, md_content, re.MULTILINE)
        data[label] = match.group(1).strip() if match else ""
    return data

def is_expired(deadline_str):
    if not deadline_str or "Ongoing" in deadline_str: return False
    try:
        # Expected format: YYYY-MM-DD
        deadline_date = datetime.strptime(deadline_str[:10], '%Y-%m-%d')
        return deadline_date < datetime.now()
    except:
        return False

def parse_tender_markdown(md_content):
    """Parses individual tender files based on the detailed template."""
    data = {}
    title_match = re.search(r'^#\s+Tender Opportunity:\s*(.+)$', md_content, re.MULTILINE)
    data['Tender Title'] = title_match.group(1).strip() if title_match else "Unknown"
    
    # Extract Category from the section below ### Category
    cat_match = re.search(r'### Category\s*\n\s*(.+)', md_content)
    data['Category'] = cat_match.group(1).strip() if cat_match else "Uncategorized"

    # Extract Tender Description from the section below ### Tender Description
    desc_match = re.search(r'### Tender Description\s*\n\s*([\s\S]+?)\n\n###', md_content)
    if not desc_match:
        desc_match = re.search(r'### Tender Description\s*\n\s*([\s\S]+?)\n\n', md_content)
    data['Tender Description'] = desc_match.group(1).strip() if desc_match else "No description provided."

    # Extract ALL Tender Documents
    docs_match = re.search(r'## Documents & Links\s*\n\s*-\s+\*\*Direct Link\*\*:\s*(.+)\n\s*-\s+\*\*Tender Documents\*\*:\s*\n([\s\S]+?)\n\n##', md_content)
    all_links = []
    if docs_match:
        direct_link = docs_match.group(1).strip()
        all_links.append(direct_link)
        # Parse individual document bullets
        doc_list = docs_match.group(2)
        bullets = re.findall(r'\[.+\]\((http[s]?://.+)\)', doc_list)
        all_links.extend(bullets)

    data['Documents / Links'] = "\n".join(list(set(all_links)))

    fields = {
        'Tender Number': r'-\s+\*\*Tender Number\*\*:\s*(.+)$',
        'Institution': r'-\s+\*\*Institution\*\*:\s*(.+)$',
        'Type': r'-\s+\*\*Tender Type\*\*:\s*(.+)$',
        'Closing Date': r'-\s+\*\*Closing Date\*\*:\s*(.+)$',
        'Status': r'-\s+\*\*Status\*\*:\s*(.+)$'
    }
    for label, pattern in fields.items():
        match = re.search(pattern, md_content, re.MULTILINE)
        data[label] = match.group(1).strip() if match else ""
    return data

def clean_filename(name):
    """Sanitizes names for Excel sheet constraints."""
    return re.sub(r'[\\/*?:\[\]]', '', name)

def create_pdf_report(df, output_path, title, header_color):
    """Generates a landscape PDF version of the report with fixed column widths."""
    doc = SimpleDocTemplate(
        str(output_path),
        pagesize=landscape(A4),
        rightMargin=20, leftMargin=20, topMargin=20, bottomMargin=20
    )
    elements = []
    styles = getSampleStyleSheet()

    # Title
    elements.append(Paragraph(f"<b>{title}</b>", styles['Title']))
    elements.append(Spacer(1, 12))

    # Prepare Data
    data = [df.columns.tolist()] + df.values.tolist()

    # Intelligent Wrapping & Column Width Calculation
    available_width = landscape(A4)[0] - 40 # Total width minus margins
    num_cols = len(df.columns)

    # Default column widths
    col_widths = [available_width / num_cols] * num_cols

    # Specific adjustments if known columns exist
    if 'Description' in df.columns:
        desc_idx = df.columns.get_loc('Description')
        col_widths[desc_idx] *= 2 # Double width for description
        # Shrink others to compensate
        remaining = available_width - col_widths[desc_idx]
        other_width = remaining / (num_cols - 1)
        for i in range(num_cols):
            if i != desc_idx: col_widths[i] = other_width

    if 'Tender Description' in df.columns:
        desc_idx = df.columns.get_loc('Tender Description')
        col_widths[desc_idx] *= 2
        remaining = available_width - col_widths[desc_idx]
        other_width = remaining / (num_cols - 1)
        for i in range(num_cols):
            if i != desc_idx: col_widths[i] = other_width

    wrapped_data = []
    style_n = styles['Normal']
    style_n.fontSize = 7
    style_n.leading = 9

    for row in data:
        wrapped_row = []
        for cell in row:
            content = str(cell).replace('\n', '<br/>')
            # Limit very long text to prevent layout overflow
            if len(content) > 2000:
                content = content[:1997] + "..."
            wrapped_row.append(Paragraph(content, style_n))
        wrapped_data.append(wrapped_row)

    t = Table(wrapped_data, colWidths=col_widths, repeatRows=1)

    # Style Table
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor(f"#{header_color}")),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('FONTSIZE', (0, 0), (-1, -1), 7),
    ]))

    elements.append(t)
    try:
        doc.build(elements)
    except Exception as e:
        print(f"⚠️ PDF build failed for {title}: {e}")

def sync_md_to_excel():
    equity_dir = Path('01_equity')
    grants_dir = Path('02_grants')
    tenders_dir = Path('03_tenders')
    pub_dir = Path('published')
    pub_dir.mkdir(parents=True, exist_ok=True)
    
    # --- 1. PROCESS EQUITY ---
    if equity_dir.exists():
        equity_dfs = []
        for md_file in equity_dir.glob('*.md'):
            if md_file.name in ['registry_audit_log.md', 'global_audit_log.md']: continue
            with open(md_file, 'r', encoding='utf-8') as f:
                df = parse_markdown_table(f.read())
                if df is not None: equity_dfs.append(df)
        
        if equity_dfs:
            full_equity = pd.concat(equity_dfs, ignore_index=True)
            
            # --- VERIFICATION FILTER ---
            # For Equity, we prioritize a 'Status' column if it exists.
            # If not, we ensure 'Source / Verification' has meaningful content.
            if 'Status' in full_equity.columns:
                full_equity = full_equity[full_equity['Status'].str.upper().isin(['ACTIVE', 'VERIFIED', ''])]
            elif 'Source / Verification' in full_equity.columns:
                # Ensure it's not just empty or "N/A"
                full_equity = full_equity[
                    full_equity['Source / Verification'].fillna('').str.len() > 3
                ]

            # Geographic Filters
            def is_sa(row):
                val = (str(row.get('Country', '')) + " " + str(row.get('Territory', ''))).lower()
                return 'south africa' in val
            
            def is_africa(row):
                val = (str(row.get('Country', '')) + " " + str(row.get('Territory', ''))).lower()
                # Simple list of major African identifiers
                african_keywords = ['africa', 'kenya', 'nigeria', 'egypt', 'ghana', 'ethiopia', 'rwanda', 'uganda']
                return any(kw in val for kw in african_keywords)

            # CLEANUP: Remove Source/Verification for published files
            cols_to_drop = ['Source', 'Source / Verification', 'Status', 'Last Verified', 'Verification Status']
            full_equity_pub = full_equity.drop(columns=[c for c in cols_to_drop if c in full_equity.columns])

            # SAVE EQUITY GLOBAL
            excel_path = pub_dir / '01_Equity_Global.xlsx'
            pdf_path = pub_dir / '01_Equity_Global.pdf'
            with pd.ExcelWriter(str(excel_path), engine='openpyxl') as writer:
                full_equity_pub.to_excel(writer, sheet_name='Global_Funders', index=False)
                apply_premium_style(writer, 'Global_Funders', INDIGO_HEADER)
            create_pdf_report(full_equity_pub, pdf_path, "Global Equity Funders", INDIGO_HEADER)
                
            # SAVE EQUITY SOUTH AFRICA
            df_sa = full_equity_pub[full_equity.apply(is_sa, axis=1)] # Filter using original for correct data
            excel_sa = pub_dir / '01_Equity_SouthAfrica.xlsx'
            pdf_sa = pub_dir / '01_Equity_SouthAfrica.pdf'
            with pd.ExcelWriter(str(excel_sa), engine='openpyxl') as writer:
                df_sa.to_excel(writer, sheet_name='SA_Focus', index=False)
                apply_premium_style(writer, 'SA_Focus', INDIGO_HEADER)
            create_pdf_report(df_sa, pdf_sa, "South Africa Focused Funders", INDIGO_HEADER)
                
            # SAVE EQUITY AFRICA
            df_africa = full_equity_pub[full_equity.apply(is_africa, axis=1)]
            excel_af = pub_dir / '01_Equity_Africa.xlsx'
            pdf_af = pub_dir / '01_Equity_Africa.pdf'
            with pd.ExcelWriter(str(excel_af), engine='openpyxl') as writer:
                df_africa.to_excel(writer, sheet_name='Africa_Focus', index=False)
                apply_premium_style(writer, 'Africa_Focus', INDIGO_HEADER)
            create_pdf_report(df_africa, pdf_af, "Africa Focused Funders", INDIGO_HEADER)

    # --- 2. PROCESS GRANTS ---
    if grants_dir.exists():
        active_grants = []
        for md_file in grants_dir.glob('*.md'):
            if md_file.name in ['template.md']: continue
            with open(md_file, 'r', encoding='utf-8') as f:
                data = parse_discrete_markdown(f.read())
                # Only include ACTIVE/VERIFIED grants that are not expired
                status = data.get('Status', '').upper()
                is_valid = status in ['ACTIVE', 'VERIFIED']
                if data and is_valid and not is_expired(data.get('Deadline', '')):
                    active_grants.append(data)
        
        if active_grants:
            df_grants = pd.DataFrame(active_grants)
            cols_to_drop = ['Source', 'Status', 'Verification Status', 'Last Verified']
            df_grants_pub = df_grants.drop(columns=[c for c in cols_to_drop if c in df_grants.columns])

            excel_gr = pub_dir / '02_Grants_Active.xlsx'
            pdf_gr = pub_dir / '02_Grants_Active.pdf'
            with pd.ExcelWriter(str(excel_gr), engine='openpyxl') as writer:
                df_grants_pub.to_excel(writer, sheet_name='Active_Grants', index=False)
                apply_premium_style(writer, 'Active_Grants', EMERALD_HEADER)
            create_pdf_report(df_grants_pub, pdf_gr, "Active Grant Opportunities", EMERALD_HEADER)

    # --- 3. PROCESS TENDERS ---
    if tenders_dir.exists():
        all_tenders = []
        for md_file in tenders_dir.glob('*.md'):
            if md_file.name in ['template.md', 'registry_audit_log.md']: continue
            with open(md_file, 'r', encoding='utf-8') as f:
                data = parse_tender_markdown(f.read())
                # Only include ACTIVE tenders
                if data and data.get('Status', '').upper() == 'ACTIVE':
                    all_tenders.append(data)
        
        if all_tenders:
            df_tenders = pd.DataFrame(all_tenders)
            # Remove status before export
            df_tenders_pub = df_tenders.drop(columns=['Status']) if 'Status' in df_tenders.columns else df_tenders

            excel_tn = pub_dir / '03_Tenders_Master.xlsx'
            pdf_tn = pub_dir / '03_Tenders_Master.pdf'

            with pd.ExcelWriter(str(excel_tn), engine='openpyxl') as writer:
                for category, group in df_tenders_pub.groupby('Category'):
                    sheet_name = clean_filename(category)[:31]
                    group.to_excel(writer, sheet_name=sheet_name, index=False)
                    apply_premium_style(writer, sheet_name, RUBY_HEADER)

            create_pdf_report(df_tenders_pub, pdf_tn, "Master Tender Registry", RUBY_HEADER)

    print("Premium Registry Expansion Complete.")

if __name__ == "__main__":
    load_environment()
    sync_md_to_excel()